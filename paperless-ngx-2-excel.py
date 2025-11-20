#!/usr/bin/env python3

import os
from urllib.parse import quote
import sys
import pwd
#from networkx import from_prufer_sequence
import requests
import inspect
import argparse
import json
import locale
import re
import zipfile
import configparser
import glob
import pprint
import asyncio
import os
import aiohttp
import re
import shutil
from time import sleep
import asyncio
import aiohttp
from datetime import datetime
from dateutil import parser
from collections import OrderedDict

from configparser import ConfigParser
from tqdm import tqdm
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from collections import OrderedDict

from datetime import datetime, timedelta
from dateutil import parser

import asyncio
import aiohttp
from pypaperless import Paperless
import glob
import os

import os
import glob

import shutil

import subprocess

import os
import subprocess
import urllib.request
import json
import re

# Neue zentrale message-Funktion + globale LOG_PATH Variable

from datetime import datetime
import os
import atexit
import unicodedata


LOG_PATH = None  # Wird beim Log-Setup gesetzt
_final_log_path = None
_last_message_was_inline = False  # Verfolgt, ob vorherige Ausgabe inline war


# --- Unicode path normalization helpers for macOS (APFS) ---
def _normalize_path(path: str) -> str:
    """Normalize Unicode for macOS (APFS often stores as NFD)."""
    try:
        import platform  # local import to avoid top-level duplication
        if platform.system() == "Darwin":
            return unicodedata.normalize("NFC", path)
    except Exception:
        pass
    return path


def safe_unlink(path: str) -> bool:
    """Remove a file or (even broken) symlink without raising outside.
    Returns True if something got removed.
    """
    path = _normalize_path(path)
    try:
        # lexists() is True for broken symlinks as well
        if os.path.lexists(path):
            try:
                os.unlink(path)
                return True
            except IsADirectoryError:
                shutil.rmtree(path)
                return True
        return False
    except FileNotFoundError:
        return False
    except PermissionError:
        # bubble up; caller should know they cannot write here
        raise
    except OSError:
        # last resort
        if os.path.exists(path):
            os.remove(path)
            return True
        return False


def message(text: str, target: str = "both", level: str = "info", inline: bool = False):
    global _last_message_was_inline

    prefix = {
        "info": "➕",
        "warn": "⚠️",
        "error": "❌"
    }.get(level, "➕")

    full_message = f"{prefix} {text}"

    # Wenn letzte Ausgabe inline war, füge einen echten Zeilenumbruch ein
    # if not inline and _last_message_was_inline:
    #     print()
    #     _last_message_was_inline = False

    if inline:
        print(full_message, end='\r', flush=True)
        _last_message_was_inline = True
        return

    if target in ("console", "both"):
        print(full_message)

    if target in ("log", "both") and LOG_PATH:
        try:
            # Kontextinfo für Log-Nachricht
            frame = inspect.currentframe().f_back
            filename = os.path.basename(frame.f_code.co_filename)
            line_number = frame.f_lineno
            function_name = frame.f_code.co_name
            context = f"{filename}:{line_number} [{function_name}]"
            with open(LOG_PATH, "a", encoding="utf-8") as f:
                f.write(f"{datetime.now()} - {context} {full_message}\n")
        except Exception as e:
            print(f"⚠️ Fehler beim Schreiben ins Log: {e}")

def set_log_path(path: str):
    global LOG_PATH
    LOG_PATH = path
    message(f"Log-Dateipfad gesetzt: {path}", target="console", level="info")

def mask_secret(secret: str, show: int = 4) -> str:
    """Maskiert einen String, zeigt nur die ersten und letzten Zeichen."""
    if not secret or len(secret) <= show * 2:
        return "*" * len(secret)
    return f"{secret[:show]}{'*' * (len(secret) - 2 * show)}{secret[-show:]}"

def xxcleanup_old_files(dir_path, filename_prefix, max_count_str, pattern="log"):
    max_count = int(max_count_str)
    glob_pattern = os.path.join(dir_path, f"{filename_prefix}*.{pattern}")
    files = sorted(glob.glob(glob_pattern), key=os.path.getmtime)
    message(f"[Cleanup] Suche: {glob_pattern} – gefunden: {len(files)} Dateien", target="log")
    if len(files) <= max_count:
        message(f"[Cleanup] Nichts zu tun: {len(files)} ≤ {max_count}", target="log")
        return
    while len(files) > max_count:
        old_file = files.pop(0)
        try:
            os.remove(old_file)
            message(f"[Cleanup] Datei gelöscht: {old_file}", target="log")
        except OSError as e:
            message(f"[Cleanup] Fehler beim Löschen von {old_file}: {e}", level="warn", target="log")

def get_log_filename(script_name, log_dir, suffix="progress"):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ext = "log" if suffix == "log" else f"{suffix}.log"
    return os.path.join(log_dir, f"##{script_name}__{timestamp}.{ext}")

def initialize_log(log_dir, script_name, max_files):
    final_log_path = get_log_filename(script_name, log_dir, "log")
    progress_log_path = get_log_filename(script_name, log_dir, "progress")
    if os.path.exists(final_log_path):
        with open(progress_log_path, "w") as new_log, open(final_log_path, "r") as old_log:
            shutil.copyfileobj(old_log, new_log)
        os.remove(final_log_path)
    else:
        open(progress_log_path, "w").close()
    cleanup_old_files(log_dir, "##" + script_name, max_files)
    return progress_log_path, final_log_path

def finalize_log():
    global LOG_PATH, _final_log_path
    if LOG_PATH and _final_log_path and os.path.exists(LOG_PATH):
        os.rename(LOG_PATH, _final_log_path)
        message("Log-Datei finalisiert.", target="console")

def prepare_logging(log_dir, script_name, max_files):
    global _final_log_path
    progress_log_path, final_log_path = initialize_log(log_dir, script_name, max_files)
    cleanup_old_files(log_dir,filename_prefix="##" + script_name ,pattern="log",max_count_str=max_files)
    _final_log_path = final_log_path
    set_log_path(progress_log_path)
    atexit.register(finalize_log)

# Beispielverwendung nach Initialisierung:
# set_log_path("pfad/zur/logdatei.log")
# message("Export gestartet", target="both")
# message("Dokument konnte nicht geladen werden", level="warn")
# message("Verarbeite Dokument 17", inline=True)


# --- Version helpers ---
def detect_git_version():
    try:
        version = subprocess.check_output(
            ["git", "describe", "--tags", "--always"],
            stderr=subprocess.DEVNULL
        ).decode("utf-8").strip()
        return version
    except Exception:
        return None

def read_version_file(path=".version"):
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                return f.read().strip()
        except:
            pass
    return None

def fallback_version():
    dt = datetime.now().strftime("%Y%m%d-%H%M")
    return f"v0.0.0-{dt}-nogit"

def get_script_version(version_file=".version"):
    """
    Build a version string based on git tag + date + short commit hash.
    If git is not available, fall back to existing .version or generated fallback.
    """
    # Try reading git tag
    try:
        tag = subprocess.check_output(
            ["git", "describe", "--tags", "--abbrev=0"],
            stderr=subprocess.DEVNULL
        ).decode("utf-8").strip()
    except Exception:
        tag = None

    # Try reading commit hash
    try:
        commit = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL
        ).decode("utf-8").strip()
    except Exception:
        commit = None

    # Always include timestamp
    dt = datetime.now().strftime("%Y%m%d-%H%M")

    if tag and commit:
        version = f"{tag}-{dt}-{commit}"
        # Create a git tag if possible
        try:
            subprocess.run(["git", "tag", "-f", version], check=False)
        except Exception:
            pass
        try:
            with open(version_file, "w") as f:
                f.write(version)
        except:
            pass
        return version

    # If git is not available, try reading existing .version
    file_version = read_version_file(version_file)
    if file_version:
        return file_version

    # Create fallback
    version = f"v0.0.0-{dt}-nogit"
    try:
        with open(version_file, "w") as f:
            f.write(version)
    except:
        pass
    return version

def get_github_repo_info():
    try:
        url = subprocess.check_output(
            ["git", "config", "--get", "remote.origin.url"],
            stderr=subprocess.DEVNULL
        ).decode("utf-8").strip()

        # z. B. https://github.com/ufe-dev/paperless-ngx-2-excel.git
        # oder git@github.com:ufe-dev/paperless-ngx-2-excel.git

        match = re.search(r"github.com[:/](.+?)/(.+?)(\.git)?$", url)
        if match:
            user, repo = match.group(1), match.group(2)
            return user, repo
    except Exception:
        pass

    return None, None

def get_github_license_identifier(user, repo):
    url = f"https://api.github.com/repos/{user}/{repo}/license"
    try:
        with urllib.request.urlopen(url) as response:
            data = json.load(response)
            return data["license"]["spdx_id"]
    except Exception as e:
        print(f"⚠️ Lizenz konnte nicht geladen werden: {e}")
        return None

def print_program_header():
    script_name = os.path.basename(__file__)
    version = get_script_version()
    user, repo = get_github_repo_info()
    license_id = get_github_license_identifier(user, repo) if user and repo else "Unbekannt"
    github_url = f"https://github.com/{user}/{repo}" if user and repo else "(kein GitHub-Repo erkannt)"

    print(f"{script_name} {version} – © 2025 {license_id} – {github_url}", end="")



def print_separator(char='#', width_ratio=2/3):
    try:
        columns = shutil.get_terminal_size().columns
    except Exception:
        columns = 80  # fallback
    print()  # Zeilenumbruch vor dem Strich
    line_width = int(columns * width_ratio)
    print(char * line_width)

def cleanup_old_files(dir_path, filename_prefix, max_count_str, pattern="log"):
    """
    Löscht alte Dateien mit bestimmtem Prefix und Endung, wenn das Limit überschritten ist.

    :param dir_path: Verzeichnis, in dem gesucht wird
    :param filename_prefix: Anfang des Dateinamens, z. B. '##steuer'
    :param max_count_str: Maximale Anzahl an Dateien als String
    :param pattern: Dateityp bzw. Dateiendung, z. B. 'log' oder 'xlsx'
    """
    max_count = int(max_count_str)
    glob_pattern = os.path.join(dir_path, f"{filename_prefix}*.{pattern}")
    files = sorted(glob.glob(glob_pattern), key=os.path.getmtime)

   # print(f"\n[Cleanup] Suche: {glob_pattern} – gefunden: {len(files)} Dateien")

    if len(files) <= max_count:
   #     print(f"[Cleanup] Nichts zu tun: {len(files)} ≤ {max_count}")
        return

    while len(files) > max_count:
        old_file = files.pop(0)
        try:
            os.remove(old_file)
            print(f"[Cleanup] Datei gelöscht: {old_file}")
        except OSError as e:
            print(f"[Cleanup] Fehler beim Löschen von {old_file}: {e}")

# ----------------------
def get_log_filename(script_name, log_dir, suffix="progress"):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if suffix == "log":
        return os.path.join(log_dir, f"##{script_name}__{timestamp}.log")
    else:
        return os.path.join(log_dir, f"##{script_name}__{timestamp}.{suffix}.log")

# ---------------------- Configuration Loading ----------------------
def load_config(config_path):
    """Lädt eine INI-Konfigurationsdatei, gibt None zurück bei Fehlern."""
    #message("process...")
    config = configparser.ConfigParser()
    try:
        config.read(config_path)
        return config
    except configparser.DuplicateSectionError as e:
        print(f"❌ Fehlerhafte INI-Datei (Duplicate Section): {config_path} – wird übersprungen. {e}")
        return None

# ----------------------
def get_script_name():
    """Return the name of the current script without extension."""
    return os.path.splitext(os.path.basename(sys.argv[0]))[0]

# ----------------------
def load_config_from_script():
    """Load the configuration from the ini file with a priority for the .ufe.ini file."""
    script_name = get_script_name()
    ufe_ini_path = f"{script_name}.ufe.ini"
    ini_path = f"{script_name}.ini"

    # Try to load the .ufe.ini file first
    if os.path.exists(ufe_ini_path):
        message(f"Using config file: {ufe_ini_path}")
        return load_config(ufe_ini_path)
    # Fallback to the .ini file
    elif os.path.exists(ini_path):
        message(f"Using config file: {ini_path}")
        return load_config(ini_path)
    else:
        print(f"Configuration files '{ufe_ini_path}' and '{ini_path}' not found.")
        sys.exit(1)

# ----------------------
def parse_currency(value):
    """Parst einen Währungswert wie 'EUR5.00' in einen Float."""
    try:
        # Entferne Währungszeichen (alles außer Ziffern, Punkt oder Minus)
        numeric_part = ''.join(c for c in value if c.isdigit() or c == '.' or c == '-')
        return float(numeric_part)
    except Exception as e:
        # print(f"Fehler beim Parsen des Währungswerts '{value}': {e}")
        return 0.0  # Fallback auf 0 bei Fehlern

# ----------------------
def format_currency(value, currency_locale="de_DE.UTF-8"):
    if value is None:
        return ""
    try:
        clean_value = ''.join(filter(str.isdigit, value))
        if not clean_value:
            return "0,00"
        value_float = float(clean_value) / 100
    except ValueError:
        value_float = 0.0

    try:
        formatted_value = locale.currency(value_float, grouping=True)
    except Exception as e:
        formatted_value = f"{value_float:.2f}"
    return formatted_value
from datetime import datetime, date

def format_date(val, output_format):
    """
    Nimmt datetime.date, datetime.datetime oder String entgegen und gibt
    'yyyy-mm' oder 'yyyy-mm-dd' zurück. Gibt None bei Fehlern.
    """
    if val is None:
        print(f"Date string is empty or None: {val}")
        return None

    # Bereits datetime/date?
    if isinstance(val, datetime):
        dt = val
    elif isinstance(val, date):
        dt = datetime(val.year, val.month, val.day)
    else:
        s = str(val).strip()
        if not s:
            print(f"Date string is empty or None: {val}")
            return None

        # Erst dd.mm.yyyy[ HH:MM], dann ISO/Varianten
        try:
            if " " in s:
                dt = datetime.strptime(s, "%d.%m.%Y %H:%M")
            else:
                dt = datetime.strptime(s, "%d.%m.%Y")
        except Exception:
            dt = None
            for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M", "%Y/%m/%d", "%d-%m-%Y", "%Y.%m.%d"):
                try:
                    dt = datetime.strptime(s, fmt)
                    break
                except ValueError:
                    continue
            if dt is None:
                print(f"Failed to format date '{val}': unsupported format")
                return None

    if output_format == "yyyy-mm":
        return dt.strftime("%Y-%m")
    if output_format == "yyyy-mm-dd":
        return dt.strftime("%Y-%m-%d")

    print(f"Unsupported output format: {output_format}")
    return None
# ----------------------
def XXformat_date(date_string, output_format):
    """
    Formatiert das Datum im Format '%d.%m.%Y' oder '%d.%m.%Y %H:%M' 
    in das gewünschte Format:
    - 'yyyy-mm' oder
    - 'yyyy-mm-dd'.
    
    Parameter:
    - date_string: Das Datum als String (im Format '%d.%m.%Y' oder '%d.%m.%Y %H:%M').
    - output_format: Das gewünschte Ausgabeformat ('yyyy-mm' oder 'yyyy-mm-dd').
    
    Rückgabe:
    - Das Datum im gewünschten Format als String oder None bei Fehlern.
    """
    if not date_string:
        print(f"Date string is empty or None: {date_string}")
        return None

    try:
        # Datum im ursprünglichen Format parsen
        if len(date_string.split(" ")) > 1:
            parsed_date = datetime.strptime(date_string, "%d.%m.%Y %H:%M")
        else:
            parsed_date = datetime.strptime(date_string, "%d.%m.%Y")
        
        # Rückgabe im gewünschten Format
        if output_format == "yyyy-mm":
            return parsed_date.strftime("%Y-%m")
        elif output_format == "yyyy-mm-dd":
            return parsed_date.strftime("%Y-%m-%d")
        else:
            print(f"Unsupported output format: {output_format}")
            return None
    except Exception as e:
        print(f"Failed to format date '{date_string}': {e}")
        return None

        return None
from datetime import datetime, date

def parse_date(val):
    # None / leere Strings
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
    else:
        s = val  # kann date/datetime/int/float sein

    # Bereits datetime/date?
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s

    # Unix-Timestamp (int/float)?
    if isinstance(s, (int, float)):
        try:
            return datetime.fromtimestamp(s).date()
        except Exception:
            pass

    # String-Parsing
    s = str(s).strip()
    # ISO zuerst
    try:
        return date.fromisoformat(s)
    except Exception:
        pass

    # alternative Formate
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d-%m-%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    # letzte Rettung: nichts parsebar
    # logger.warning(f"[parse_date] Could not parse {s!r}")
    return None
# ----------------------

def _default_retry_exceptions():
    import aiohttp, asyncio
    excs = [aiohttp.ClientError, asyncio.TimeoutError, ConnectionResetError, BrokenPipeError]
    # Aiohttp: ServerDisconnectedError ist ein eigener Typ -> explizit mit rein
    try:
        from aiohttp.client_exceptions import ServerDisconnectedError
        excs.append(ServerDisconnectedError)
    except Exception:
        pass
    # pypaperless-Wrapper
    try:
        from pypaperless.exceptions import PaperlessConnectionError
        excs.append(PaperlessConnectionError)
    except Exception:
        pass
    return tuple(excs)

async def retry_async(fn, retries=5, delay=2, backoff=2.0, jitter=0.3,
                      exceptions=None, desc=None):
    """
    Führt fn() mit Exponential Backoff + leichtem Jitter aus.
    - exceptions: Tuple der Exceptions, die zu Retries führen (Default: Netzwerk + PaperlessConnectionError)
    """
    import random, shutil as _shutil, asyncio as _asyncio
    if exceptions is None:
        exceptions = _default_retry_exceptions()

    current_delay = float(delay)
    for attempt in range(1, retries + 1):
        try:
            return await fn()
        except exceptions as e:
            if attempt == retries:
                raise
            label = f' bei "{desc}"' if desc else ''
            term_width = _shutil.get_terminal_size((80, 20)).columns
            print(f"\r[retry_async] Fehler{label}: {e} – Versuch {attempt}/{retries}, nächster in {current_delay:.1f}s...".ljust(term_width), end='', flush=True)
            # kleiner Jitter, damit viele gleichzeitige Requests nicht synchron wieder zuschlagen
            wait = current_delay * (1.0 + random.uniform(-jitter, jitter))
            await _asyncio.sleep(max(0.1, wait))
            current_delay *= backoff

# ----------------------
def should_export(export_dir: str, frequency: str, config_mtime: float) -> tuple[bool, str]:

    base_name = os.path.basename(export_dir)
    latest_xlsx_mtime = None

    for fname in os.listdir(export_dir):
        if fname.startswith(f"##{base_name}-") and fname.endswith(".xlsx"):
            fpath = os.path.join(export_dir, fname)
            mtime = os.path.getmtime(fpath)
            if latest_xlsx_mtime is None or mtime > latest_xlsx_mtime:
                latest_xlsx_mtime = mtime

    if latest_xlsx_mtime is None:
        return True, "Keine .xlsx-Datei vorhanden"

    if config_mtime > latest_xlsx_mtime:
        config_time = datetime.fromtimestamp(config_mtime).strftime('%Y-%m-%d %H:%M:%S')
        xlsx_time = datetime.fromtimestamp(latest_xlsx_mtime).strftime('%Y-%m-%d %H:%M:%S')
        return True, (
            f"Config modified: "
            f"(INI: {config_time}, XLSX: {xlsx_time})"
        )
    
    now = datetime.now()
    last_export = datetime.fromtimestamp(latest_xlsx_mtime)
    frequency = frequency.lower().strip()

    readable_time = last_export.strftime('%Y-%m-%d %H:%M:%S')
    # Bedingung beschreiben
    if frequency == "hourly":
        next_export = last_export + timedelta(hours=1)
        if now > next_export:
            return True, f"hourly: last={readable_time}, next={next_export}, now={now}"
        else:
            return False, f"noExport: hourly until {next_export.strftime('%Y-%m-%d %H:%M:%S')}"
    
    elif frequency == "4hourly":
        next_export = last_export + timedelta(hours=4)
        if now > next_export:
            return True, f"4hourly: last={readable_time}, next={next_export}, now={now}"
        else:
            return False, f"noExport: 4hourly until {next_export.strftime('%Y-%m-%d %H:%M:%S')}"

    elif frequency in ("daily", "weekday"):
        next_export = last_export + timedelta(days=1)
        if now > next_export:
            return True, f"daily: last={readable_time}, next={next_export}, now={now}"
        else:
            return False, f"noExport: daily until {next_export.strftime('%Y-%m-%d %H:%M:%S')}"
    
    elif frequency == "weekly":
        next_export = last_export + timedelta(days=7)
        if now > next_export:
            return True, f"weekly: last={readable_time}, next={next_export}, now={now}"
        else:
            return False, f"noExport: weekly until {next_export.strftime('%Y-%m-%d')}"
    
    elif frequency == "monthly":
        next_month = (last_export.replace(day=1) + timedelta(days=32)).replace(day=1)
        if now > next_month:
            return True, f"monthly: last={last_export.strftime('%Y-%m')}, next={next_month.strftime('%Y-%m')}, now={now.strftime('%Y-%m-%d')}"
        else:
            return False, f"noExport: monthly until {next_month.strftime('%Y-%m-%d')}"

    elif frequency == "yearly":
        next_year = datetime(last_export.year + 1, 1, 1)
        if now > next_year:
            return True, f"yearly: last={last_export.year}, next={next_year.year}, now={now.year}"
        else:
            return False, f"noExport: yearly until {next_year.year}"
    return False, (
        f"noExport: last file {readable_time}, "
    )

async def get_dict_from_paperless(endpoint):
    """
    Generische Funktion, um ein Dictionary aus einem Paperless-Endpoint zu erstellen.
    Erwartet ein `endpoint`-Objekt, das eine `all()`-Methode und einen Abruf per ID unterstützt.
    """
    items = await retry_async(fn=lambda: endpoint.all())
    #items = await endpoint.all()
    item_dict = {}

    for itemKey in items:
        item = await endpoint(itemKey)
        item_dict[item.id] = item  # Speichert das gesamte Objekt mit der ID als Schlüssel

    return item_dict  # Gibt ein Dictionary {ID: Objekt} zurück
# Modulweiter Cache (z. B. ganz oben im Script)
_paperless_meta_cache = None

async def fetch_paperless_meta(paperless, force_reload=False):
    global _paperless_meta_cache

    if _paperless_meta_cache is not None and not force_reload:
        return _paperless_meta_cache

    def log_and_print(name):
        message(text=f"getting {name}...",inline=True)

    meta = {}

    for name, endpoint in {
        "storage_paths": paperless.storage_paths,
        "correspondents": paperless.correspondents,
        "document_types": paperless.document_types,
        "tags": paperless.tags,
        "users": paperless.users,
        "custom_fields": paperless.custom_fields
    }.items():
        message(f"{name}",inline=True)
        try:
            meta[name] = await get_dict_from_paperless(endpoint)
            message(f"{name.capitalize()}: {len(meta[name])}")
        except Exception as e:
            message( f"Fehler beim Abrufen von {name}: {e}")
            meta[name] = {}  # Leere Liste als Fallback, damit getmeta nicht crasht

    _paperless_meta_cache = meta
    return meta

# ----------------------
def getmeta(key, doc, meta):
    """
    Liefert einen aufgelösten Namen (oder 'Unbekannt') für Felder wie:
      - 'correspondent'  -> meta['correspondents'][id].name
      - 'document_type'  -> meta['document_types'][id].name
      - 'storage_path'   -> meta['storage_paths'][id].name
      - 'tags'           -> Liste von IDs -> kommagetrennte Namen
    Arbeitet robust mit Dict-Metadaten und None/fehlenden IDs.
    """
    try:
        value = getattr(doc, key, None)

        # Tags: Liste von IDs -> kommagetrennte Namen
        if key == "tags":
            if not isinstance(value, list):
                return "Unbekannt"
            tags_map = meta.get("tags") or {}
            names = []
            for tid in value:
                obj = tags_map.get(tid)
                name = getattr(obj, "name", None) if obj else None
                if name:
                    names.append(name)
            return ", ".join(names) if names else "Unbekannt"

        # Einzelfelder: ID -> Objekt -> .name
        space = f"{key}s"  # correspondent -> correspondents, document_type -> document_types, storage_path -> storage_paths
        space_map = meta.get(space) or {}
        if value is None:
            return "Unbekannt"
        obj = space_map.get(value)
        if not obj:
            return "Unbekannt"
        return getattr(obj, "name", None) or getattr(obj, "username", "Unbekannt")
    except Exception as e:
        print(f"Fehler beim Auflösen von '{key}': {e}")
        return "Unbekannt"

async def export_pdf(doc, working_dir):
    """Exportiert ein Dokument als PDF mit Retry; bei endgültigem Fehler wird geloggt und übersprungen."""
    sanitized_title = sanitize_filename(doc.title)
    filename = f"{doc.id}--{sanitized_title}.pdf"
    pdf_path = os.path.join(working_dir, filename)

    try:
        download = await retry_async(lambda: doc.get_download(),
                                     desc=f"PDF-Download für Dokument {doc.id}")
        document_content = download.content
        if not document_content:
            message(f"Keine PDF-Daten für Dokument {doc.id} gefunden.", target="log", level="warn")
            return False

        with open(pdf_path, 'wb') as f:
            f.write(document_content)
        return True
    except Exception as e:
        message(f"PDF-Download fehlgeschlagen für Doc {doc.id}: {e}", target="log", level="warn")
        return False

# ----------------------

def sanitize_filename(filename):
    """
    Remove or replace characters in the filename that are problematic in common filesystems.
    """
    original = filename

    # Ersetze bekannte problematische Zeichen (Windows, Samba, NAS, etc.)
    sanitized = re.sub(r'[<>:"/\\|?*\[\]]', '-', filename)  # auch [ und ] ersetzen

    # Optional: Leerzeichen am Anfang/Ende oder doppelte Minus entfernen
    sanitized = re.sub(r'\s+', ' ', sanitized).strip()
    sanitized = re.sub(r'-{2,}', '-', sanitized)

    # Truncate auf maximale Pfadlänge
    sanitized = sanitized[:255]

  #  if sanitized != original:
  #      message(f"Dateiname bereinigt: '{original}' → '{sanitized}'", target="both")

    return sanitized

# Remove characters not allowed by Excel (control chars 0x00-0x1F except tab/newline/carriage return)
_XL_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

def clean_for_excel(value):
    """Return value with illegal worksheet characters removed (strings only)."""
    if value is None:
        return None
    if isinstance(value, str):
        return _XL_ILLEGAL_RE.sub('', value)
    return value

# ----------------------

def get_document_json(paperless,doc):
    api_token = paperless._token # Dein-Token
    headers = {"Authorization": f"Token {api_token}"}

    path=doc._api_path 
    url=paperless._base_url

    """Retrieve detailed document metadata from Paperless API."""
    response = requests.get(f"{url}/{path}", headers=headers)
    
    if response.status_code == 200:
        return response.json()  # Die JSON-Daten des Dokuments zurückgeben
    else:
        raise Exception(f"Failed to fetch document metadata: {response.status_code}")

def export_json(paperless, doc, working_dir):
    """Export a document's metadata as JSON."""
    sanitized_title = sanitize_filename(doc.title)
    filename = f"{doc.id}--{sanitized_title}.json"
    json_path = os.path.join(working_dir, filename)

    detailed_doc = get_document_json(paperless=paperless, doc=doc)
    with open(json_path, "w", encoding="utf-8") as json_file:
        json.dump(detailed_doc, json_file, ensure_ascii=False, indent=4)

# ---------------------- Excel Export Helpers ----------------------
# ----------------------
def debug_write(df, directory):
    print("🔧 DEBUG: Isoliere XLSX-Erstellung…")
    # pandas removed – debug stubs disabled
    pass

def export_to_excel(data, file_path, script_name, currency_columns, dir, url, meta, maxfiles, query, frequency):
    import re, os, shutil
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
    import pwd

    # --- COLLECT ALL HEADERS ACROSS ALL DOCUMENTS (global header set) ---
    all_headers = []
    header_set = set()
    for row in data:
        for k in row.keys():
            if k not in header_set:
                header_set.add(k)
                all_headers.append(k)

    # Normalize all rows: ensure each row has all headers (missing -> None)
    normalized_data = []
    for row in data:
        fixed = {}
        for h in all_headers:
            fixed[h] = row.get(h, None)
        normalized_data.append(fixed)

    data = normalized_data

    directory = os.path.dirname(file_path)
    base_dirname = os.path.basename(directory)
    today = datetime.now().strftime("%Y%m%d")

    history_prefix = f"##{base_dirname}-{today}"
    existing = []
    for f in os.listdir(directory):
        if f.startswith(history_prefix) and f.endswith(".xlsx"):
            try:
                n = int(f.rsplit("-", 1)[1].split(".")[0])
                existing.append(n)
            except:
                pass

    next_num = 0 if not existing else max(existing) + 1
    history_filename = f"{history_prefix}-{next_num}.xlsx"
    fullfilename = os.path.join(directory, history_filename)

    static_filename = os.path.join(directory, f"##{base_dirname}.xlsx")

    wb = Workbook()
    # --- ALWAYS KEEP AT LEAST ONE SHEET ---
    ws_default = wb.active

    if not data:
        ws_default.title = "Keine_Daten"
        ws_default.cell(row=1, column=1, value="Keine Dokumente gefunden.")
        wb.save(fullfilename)
        return

    # If there IS data → replace default sheet
    wb.remove(ws_default)

    headers = all_headers

    # Sheet1 disabled — no data written here
    row_idx = 4

    # --- UNFORMATTED TABLE SHEET (A1) ---
    ws_plain = wb.create_sheet("Tabelle")

    # Write headers at row 1
    for col_idx, col_name in enumerate(headers, start=1):
        ws_plain.cell(row=1, column=col_idx, value=col_name)

    # Write data starting at row 2
    plain_row = 2
    for row in data:
        for col_idx, col_name in enumerate(headers, start=1):
            ws_plain.cell(row=plain_row, column=col_idx, value=row[col_name])
        plain_row += 1

    from openpyxl.utils import get_column_letter
    last_plain_row = plain_row - 1
    last_plain_col = get_column_letter(len(headers))
    plain_ref = f"A1:{last_plain_col}{last_plain_row}"

    # --- Add working hyperlinks in LINK column ---
    paperless_url = str(url).rstrip("/")
    link_col_index = None

    # find LINK column index
    for idx, h in enumerate(headers, start=1):
        if h == "LINK":
            link_col_index = idx
            break

    if link_col_index:
        for r in range(2, last_plain_row + 1):
            doc_id = ws_plain.cell(row=r, column=1).value
            link_cell = ws_plain.cell(row=r, column=link_col_index)
            link_cell.value = doc_id
            link_cell.hyperlink = f"{paperless_url}/documents/{doc_id}"
            link_cell.style = "Hyperlink"

    # --- UNFORMATTED TABLE: CREATE DATA TABLE WITH STYLE ---
    import re
    raw_tbl = f"{base_dirname}"
    safe_tbl = re.sub(r'[^A-Za-z0-9]', '', raw_tbl)
    if not safe_tbl:
        safe_tbl = "Data"
    table_name = f"tbl{safe_tbl}"
    data_table = Table(displayName=table_name, ref=plain_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    data_table.tableStyleInfo = style
    ws_plain.add_table(data_table)

    message("[DEBUG] Unformatted table sheet created", target="both")

    # debug_path3 = os.path.join(directory, f"##{base_dirname}-state3.xlsx")
    # wb.save(debug_path3)

    # --- EARLY MAIN TABLE DISABLED (last_col_letter not yet defined) ---
    # (moved table creation to the correct position after last_col_letter is known)

    last_data_row = row_idx - 1
    last_col_letter = get_column_letter(len(headers))

    # Sheet1 autosize disabled

    # --- Enhanced Metadata Sheet ---
    ws_meta = wb.create_sheet("📊 Metadaten")

    rows = []
    r = rows.append

    # HeaderInfo into metadata sheet
    r(["HeaderInfo", f"{script_name} -- {directory} -- {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} -- {pwd.getpwuid(os.getuid()).pw_name} -- {os.uname().nodename}"])
    r(["", ""])
    # 📝 Exportinformationen
    r(["📝 Exportinformationen", ""])
    from inspect import getsourcefile
    r(["Script-Version", get_script_version()])
    r(["Export-Datum", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    r(["Hostname", os.uname().nodename])
    r(["Username", pwd.getpwuid(os.getuid()).pw_name])
    r(["", ""])

    # 📁 Verzeichnisse & Dateigrößen
    xlsx_size = os.path.getsize(fullfilename) if os.path.exists(fullfilename) else 0
    r(["📁 Verzeichnisse & Dateigrößen", ""])
    r(["Export-Verzeichnis", directory])
    r(["Excel-Datei", os.path.basename(fullfilename)])
    r(["Größe (xlsx)", f"{xlsx_size/1024:.2f} KB"])
    r(["JSON-Dateien", len([f for f in os.listdir(directory) if f.endswith('.json')])])
    r(["PDF-Dateien", len([f for f in os.listdir(directory) if f.endswith('.pdf')])])
    r(["", ""])

    # ⚙️ Config
    r(["⚙️ Konfiguration (config.ini)", ""])
    r(["Query", query])
    # Encode query strictly so that spaces, colons, parentheses, etc. don't break Paperless search
    safe_query = quote(str(query), safe="")
    r(["Query-Link", f"{paperless_url}/documents/?q={safe_query}"])
    r(["API-Query-Link", f"{paperless_url}/api/documents/?query={query}"])
    r(["API-Query-Link-Raw", f"{paperless_url}/api/documents/?query={query}"])
    r(["Frequency", frequency])
    r(["", ""])

    # 📊 Dokument-Statistik
    r(["📊 Dokument-Statistik", ""])
    r(["Dokumente gesamt", len(data)])
    r(["Currency-Felder", ", ".join(currency_columns)])
    r(["Header-Spalten", ", ".join(headers)])
    r(["", ""])

    # 🧩 Custom Fields
    r(["🧩 Custom Fields", ""])
    all_custom_fields = [c for c in headers if c not in ("ID","LINK","Korrespondent","Titel","Tags","Seiten","Dokumenttyp","Speicherpfad")]
    r(["Anzahl Custom Fields", len(all_custom_fields)])
    r(["Custom Fields", ", ".join(all_custom_fields)])
    r(["", ""])

    # 🐍 Python-Umgebung
    import importlib.metadata
    installed_packages = sorted([f"{dist.metadata['Name']}=={dist.version}" for dist in importlib.metadata.distributions()])
    r(["🐍 Python Umgebung", ""])
    r(["Python-Version", platform.python_version()])
    r(["Pakete (Top 10)", ", ".join(installed_packages[:10])])
    r(["", ""])

    # Schreibe alles in das Sheet
    for row_idx, (colA, colB) in enumerate(rows, start=1):
        ws_meta.cell(row=row_idx, column=1, value=colA)
        ws_meta.cell(row=row_idx, column=2, value=colB)

    # --- Convert Query-Link row into real hyperlink + keep raw text ---
    for row_idx, (colA, colB) in enumerate(rows, start=1):
        if colA == "Query-Link":
            cell = ws_meta.cell(row=row_idx, column=2)
            cell.hyperlink = colB
            cell.style = "Hyperlink"
            # Add raw text version one row below
            ws_meta.insert_rows(row_idx + 1)
            ws_meta.cell(row=row_idx + 1, column=1, value="Query-Link-Text")
            ws_meta.cell(row=row_idx + 1, column=2, value=colB)
            break

    # --- Style metadata sheet ---
    from openpyxl.styles import Alignment, Border, Side

    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, (colA, colB) in enumerate(rows, start=1):
        cell_a = ws_meta.cell(row=row_idx, column=1)
        cell_b = ws_meta.cell(row=row_idx, column=2)
        cell_a.alignment = Alignment(vertical="top")
        cell_b.alignment = Alignment(vertical="top")
        cell_a.border = border
        cell_b.border = border

    # style section headers
    for row_idx, (colA, colB) in enumerate(rows, start=1):
        if colB == "":
            ws_meta.cell(row=row_idx, column=1).font = Font(bold=True, size=13, color="1F4E79")

    # debug_path5 = os.path.join(directory, f"##{base_dirname}-state5.xlsx")
    # wb.save(debug_path5)
    wb.save(fullfilename)

    try:
        if os.path.exists(static_filename) or os.path.islink(static_filename):
            safe_unlink(static_filename)
        shutil.copy2(fullfilename, static_filename)
    except Exception as e:
        message(f"⚠️ Fehler beim Erstellen der statischen Datei: {e}", "both")

    message(f"Excel-Datei erfolgreich erstellt: {fullfilename}")

# ----------------------
def has_file_from_today(directory):
    """
    Prüft, ob im angegebenen Verzeichnis eine Datei existiert,
    die heute erstellt oder zuletzt geändert wurde.
    """
    today = datetime.now().date()
    if not os.path.exists(directory):
        return False

    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if os.path.isfile(file_path):
            # Änderungszeitpunkt der Datei
            file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
            if file_mtime.date() == today:
                return True
    return False
def process_custom_fields(meta, doc):
    """
    - meta['custom_fields']: Dict[id -> Objekt], Objekt._data enthält u.a. 'data_type' und evtl. 'extra_data'
    - doc['custom_fields']: Liste von Einträgen mit 'field' (ID) und 'value'

    Rückgabe:
      custom_fields: dict mit aufgelösten Werten/Labels
      currency_fields: Liste von Feldnamen (für Summen)
    """
    custom_fields = {}
    currency_fields = []

    meta_cf = meta.get("custom_fields") or {}
    doc_cf_list = doc.get("custom_fields") or []

    for entry in doc_cf_list:
        field_id = entry.get("field")
        if not field_id or field_id not in meta_cf:
            continue

        meta_obj = meta_cf[field_id]
        field_name = getattr(meta_obj, "name", f"field_{field_id}")

        data = getattr(meta_obj, "_data", {}) or {}
        field_type = (data.get("data_type") or "").lower()
        field_value = entry.get("value")

        # Choices/Options einsammeln (Liste oder Dict → Mapping)
        extra = (data.get("extra_data") or {})
        choices = extra.get("select_options") or extra.get("choices") or extra.get("options") or []

        choice_map = {}
        if isinstance(choices, dict):
            choice_map = {str(k): v for k, v in choices.items()}
        elif isinstance(choices, list):
            for ch in choices:
                if isinstance(ch, dict):
                    key = ch.get("value") or ch.get("id") or ch.get("key") or ch.get("slug") or ch.get("label")
                    label = ch.get("label") or ch.get("name") or ch.get("value") or ch.get("slug") or str(ch)
                else:
                    key = str(ch)
                    label = str(ch)
                if key is not None:
                    choice_map[str(key)] = label

        def resolve_choice(v):
            if v is None:
                return "none"
            if isinstance(v, list):  # multiselect
                return ", ".join(str(choice_map.get(str(x), x)) for x in v)
            return choice_map.get(str(v), v)

        if field_type in ("monetary", "currency"):
            numeric_value = parse_currency(field_value)  # deine bestehende Funktion
            custom_fields[field_name] = numeric_value
            custom_fields[f"{field_name}_formatted"] = format_currency(field_value)  # deine bestehende Funktion
            currency_fields.append(field_name)

        elif field_type in ("select", "multiselect", "choice", "choices"):
            custom_fields[field_name] = resolve_choice(field_value)

        else:
            custom_fields[field_name] = field_value

    return custom_fields, currency_fields
# ----------------------
def XXprocess_custom_fields(meta, doc):
    custom_fields = {}
    currency_fields = []  # Liste zum Speichern der Currency-Feldnamen

    if "custom_fields" in doc:
        for custom_field in doc["custom_fields"]:
            field_id = custom_field.get("field")
            if not field_id:    
                continue
            field_name = meta["custom_fields"][field_id].name
            field_type = meta['custom_fields'][field_id]._data['data_type']
            field_value = custom_field.get("value")
            
            if field_type == "monetary":
                numeric_value = parse_currency(field_value)
                custom_fields[field_name] = numeric_value  # Rohdaten speichern
                custom_fields[f"{field_name}_formatted"] = format_currency(field_value)  # Formatierte Version speichern
                currency_fields.append(field_name)  # Speichern des Currency-Felds
            elif field_type == "select":
                # Hole die choices aus meta["custom_fields"][field_id] und prüfe, ob field_value None ist
                choices = meta['custom_fields'][field_id]._data['extra_data']['select_options']
                if field_value is None:
                    custom_fields[field_name] = "none"  # Wenn field_value None ist, setze "none"
                else:
                    # Wenn field_value nicht None ist, hole den Wert aus den choices
                    custom_fields[field_name] =  choices[field_value]

            else:
                custom_fields[field_name] = field_value

    return custom_fields, currency_fields

async def get_documents_with_retry(paperless, query):
    return await retry_async(
        lambda: collect_async_iter(paperless.documents.search(query)),
        desc=f"Dokumente für Query '{query}'"
    )

async def collect_async_iter(aiter):
    return [item async for item in aiter]

async def search_documents(paperless, query):
    return [item async for item in paperless.documents.search(query)]

def find_cached_file(doc_id, all_dir, kind):
    """Findet Datei im .all-Verzeichnis anhand von doc_id und Dateityp ('pdf' oder 'json')"""
    prefix = f"{doc_id}--"
    suffix = f".{kind}"
    for fname in os.listdir(all_dir):
        if fname.startswith(prefix) and fname.endswith(suffix):
            return os.path.join(all_dir, fname)
    return None

import os
import shutil
import platform

def is_synology():
    uname = platform.uname()
    return (
        "synology" in uname.node.lower()
        or os.path.exists("/etc/synoinfo.conf")
        or os.path.exists("/etc.defaults/synoinfo.conf")
    )

def force_copy_mode():
    """
    Gibt True zurück, wenn per Umgebungsvariable 'FORCE_COPY=1' erzwungen wird,
    dass keine Symlinks oder Hardlinks verwendet werden sollen.
    """
    return os.environ.get("FORCE_COPY", "0") == "1"

import os
import shutil
import glob

def link_export_file(doc, kind, working_dir, all_dir=".all"):
    assert kind in ("pdf", "json")

    filename = f"{doc.id}--{sanitize_filename(doc.title)}.{kind}"
    dest_path = os.path.join(working_dir, filename)
    dest_dir = os.path.dirname(dest_path)

    # Normalize for macOS Unicode edge cases
    dest_path = _normalize_path(dest_path)
    dest_dir = _normalize_path(dest_dir)

    # Quelle im .all-Ordner finden
    src_path = find_cached_file(doc.id, all_dir=all_dir, kind=kind)
    if src_path is None:
        raise FileNotFoundError(f"Keine {kind.upper()}-Datei für Dokument {doc.id} im .all-Verzeichnis gefunden")
    src_path = _normalize_path(src_path)

    os.makedirs(dest_dir, exist_ok=True)

    # Alte Dateien mit gleicher doc.id löschen, wenn Name abweicht
    prefix = f"{doc.id}--"
    pattern = os.path.join(dest_dir, f"{prefix}*.{kind}")
    existing_files = glob.glob(pattern)

    for old_path in existing_files:
        if os.path.abspath(old_path) != os.path.abspath(dest_path):
            try:
                safe_unlink(old_path)
            except Exception as e:
                message(f"⚠️ Fehler beim Entfernen alter Datei: {old_path} → {e}", "both")

    # Prüfen ob Ziel existiert oder kaputter Symlink vorhanden ist
    if os.path.lexists(dest_path):
        try:
            if os.path.islink(dest_path) and os.path.realpath(dest_path) == os.path.realpath(src_path):
                return "symlink (OK)"
            elif os.path.exists(dest_path) and os.path.samefile(dest_path, src_path):
                return "hardlink/copy (OK)"
            else:
                safe_unlink(dest_path)
        except Exception as e:
            message(f"Fehler beim Entfernen von bestehender Datei: {e}", "both")
            safe_unlink(dest_path)

    # 🔁 Nur kopieren, wenn Umgebungsvariable gesetzt ist
    if force_copy_mode():
        try:
            shutil.copy2(src_path, dest_path)
            if os.path.exists(dest_path) and os.path.getsize(dest_path) > 0:
                return "copy (FORCE)"
        except Exception as e:
            raise RuntimeError(f"Konnte Datei nicht kopieren: {src_path}")

    # 🔗 Symlink versuchen (relativer Pfad!)
    try:
        rel_src_path = os.path.relpath(src_path, start=dest_dir)
        os.symlink(rel_src_path, dest_path)
        if os.path.exists(dest_path) and os.path.getsize(dest_path) > 0:
            return "symlink (neu)"
    except Exception as e:
        message(f"Symlink fehlgeschlagen: {type(e).__name__}: {e}", "both")
        message(f"                      : {rel_src_path}", "both")
        message(f"                      : {dest_path}", "both")

    # 🔗 Hardlink versuchen
    try:
        os.link(src_path, dest_path)
        if os.path.exists(dest_path) and os.path.getsize(dest_path) > 0:
            return "hardlink (neu)"
    except Exception as e:
        message(f"Hardlink fehlgeschlagen: {e}", "both")

    # Letzter Fallback: Kopieren
    try:
        shutil.copy2(src_path, dest_path)
        if os.path.exists(dest_path) and os.path.getsize(dest_path) > 0:
            return "copy (fallback)"
    except Exception as e:
        message(f"Kopie fehlgeschlagen: {e}", "both")

    raise RuntimeError(f"Zieldatei konnte nicht erstellt werden: {dest_path}")

async def exportThem(paperless, dir, query, max_files, frequency, api_url):
    count = 0 
    """Process and export documents"""
    document_data = []
    currency_columns = []  # Liste zur Speicherung aller Currency-Felder
    custom_fields = {}
    meta = await fetch_paperless_meta(paperless)

#    documents = [item async for item in paperless.documents.search(query)]
    documents = await retry_async(
       lambda: search_documents(paperless, query),
       desc=f"Dokumente für Query '{query}'"
       )

    # --- HARD CLEAN: remove all existing PDF/JSON files in this export directory ---
    try:
        removed = 0
        for fname in os.listdir(dir):
            if fname.endswith(".pdf") or fname.endswith(".json"):
                fpath = os.path.join(dir, fname)
                try:
                    os.remove(fpath)
                    removed += 1
                except Exception as e:
                    message(f"⚠️ Fehler beim Löschen {fpath}: {e}", target="log")
        message(f"🧹 HARD CLEAN: {removed} Datei(en) aus {dir} gelöscht.", target="both")
    except Exception as e:
        message(f"❌ Fehler beim HARD CLEAN in {dir}: {e}", target="both")

    #documents = await retry_async(
    #   lambda: collect_async_iter(paperless.documents.search(query)),
    #    desc=f"Dokumente für Query '{query}'"
    #)

    for doc in tqdm(documents, desc=f"Processing documents for '{dir} with {query}'", unit="doc"):
        count += 1

        # Metadaten optional abrufen (standardmäßig AUS, wegen pypaperless 5.x Validierung)
        if os.environ.get("FETCH_METADATA", "0") == "1":
            try:
                _ = await retry_async(lambda: doc.get_metadata(), desc=f"Metadaten für Dokument {doc.id}")
            except Exception as e:
                message(f"Metadaten für Dokument {doc.id} übersprungen: {e}", target="log")

        docData = doc._data
        page_count = docData['page_count']
        message(f"{doc.id} {doc.title} {page_count} pages",target="log")

        custom_fields, doc_currency_columns = process_custom_fields(meta=meta,doc=docData)
        currency_columns.extend(doc_currency_columns)  # Speichere Currency-Felder
        thisTags =  getmeta("tags", doc, meta=meta)

        # --- Normalize Storage Path into a constant prefix and a variable template tail ---
        sp_obj = meta["storage_paths"].get(doc.storage_path)
        sp_name = getattr(sp_obj, "name", "") if sp_obj else ""
        # Paperless storage path objects expose the jinja-like path template commonly as "path"
        sp_template = ""
        if sp_obj:
            sp_template = getattr(sp_obj, "path", "") or getattr(getattr(sp_obj, "_data", {}), "get", lambda *_: "")("path") if hasattr(sp_obj, "_data") else ""

        # Extract constant prefix (everything before the first '{{') and the variable tail (starting with '{{')
        _tpl = sp_template or ""
        if "{{" in _tpl:
            _pre, _tail = _tpl.split("{{", 1)
            sp_prefix = _pre.rstrip("/ ").strip()
            sp_tail = "{{" + _tail
        else:
            sp_prefix = _tpl.rstrip("/ ").strip()
            sp_tail = ""

        # A compact grouping key that ignores location-specific prefix and groups by the shared template
        # (so that variants like "GutHembach St/<same template>" and "AndererOrt/<same template>" collapse)
        storage_group = sp_tail.strip() or sp_name or sp_prefix

        # Daten für die Excel-Tabelle sammeln
        row = OrderedDict([
            ("ID", doc.id),
            ("LINK", doc.id),
            ("Korrespondent", getmeta("correspondent", doc, meta)),
            ("Titel", doc.title),
            ("Tags", thisTags),
          #  ("StoragePathName", sp_name),
          #  ("StoragePrefix", sp_prefix),
          #  ("StorageGroup", storage_group),
            # Custom Fields direkt hinter den Tags Pieinfügen
            *custom_fields.items(),
            ("Seiten", doc._data['page_count']),
            ("Dokumenttyp", getmeta("document_type", doc, meta)),
            ("Speicherpfad", getmeta("storage_path", doc, meta)),
            ("ArchivDate", parse_date(doc.created)),
            ("ArchivedDateMonth", format_date(parse_date(doc.created), "yyyy-mm")),
            ("ArchivedDateFull", format_date(parse_date(doc.created), "yyyy-mm-dd")),
            ("ModifyDate", parse_date(doc.modified)),
            ("ModifyDateMonth", format_date(parse_date(doc.modified), "yyyy-mm")),
            ("ModifyDateFull", format_date(parse_date(doc.modified), "yyyy-mm-dd")),
            ("AddedDate", parse_date(doc.added)),
            ("AddDateMonth", format_date(parse_date(doc.added), "yyyy-mm")),
            ("AddDateFull", format_date(parse_date(doc.added), "yyyy-mm-dd")),
            ("OriginalName", doc.original_file_name),
            ("ArchivedName", doc.archived_file_name),
            ("Owner", getattr(meta["users"].get(doc.owner), "username", "Unbekannt") if doc.owner else "Unbekannt"),
            ("URL", f"{api_url}/documents/{doc.id}")
        ]
        )

        document_data.append(row)

        # Exportiere das PDF des Dokuments
        #await export_pdf(doc, working_dir=dir)
        #export_json(paperless=paperless,doc=doc,working_dir=dir)
        # Statt export_pdf / export_json:
        export_dir  = os.path.dirname(dir)

        try:
            method_pdf = link_export_file(doc, kind="pdf", working_dir=dir, all_dir=os.path.join(export_dir, ".all"))
        except Exception as e:
            method_pdf = "ERROR"
            message(f"❌ PDF fehlt für Doc {doc.id}: {e}", target="both", level="warn")

        try:
            method_json = link_export_file(doc, kind="json", working_dir=dir, all_dir=os.path.join(export_dir, ".all"))
        except Exception as e:
            method_json = "ERROR"
            message(f"❌ JSON fehlt für Doc {doc.id}: {e}", target="both", level="warn")
        #method_pdf = link_export_file(doc, kind="pdf", working_dir=dir, all_dir=os.path.join(export_dir, ".all")) 
        #method_json = link_export_file(doc, kind="json", working_dir=dir, all_dir=os.path.join(export_dir, ".all"))

       # message(f"{doc.id}: PDF → {method_pdf} json → {method_json}", target="both")


    # Exportiere die gesammelten Daten nach Excel
    #path=doc._api_path 
    url=paperless._base_url

    last_dir = os.path.basename(dir)
    # Use the query for the Excel filename, but sanitize it to remove illegal characters (e.g., '*', '?', etc.).
    #safe_query = sanitize_filename(str(query) if query is not None else "")
    #if not safe_query:
        # Fallback to directory name if query is empty after sanitization
    safe_query = sanitize_filename(last_dir)

    excel_file = os.path.join(dir, f"##{safe_query}-{datetime.now().strftime('%Y%m%d')}.xlsx")
    #export_to_excel(document_data, excel_file, get_script_name, currency_columns=currency_columns,dir=dir, url=url,meta=meta, maxfiles=max_files,query=query, frequency=frequency)
    export_to_excel(document_data, excel_file, get_script_name(), currency_columns=currency_columns, dir=dir, url=url, meta=meta, maxfiles=max_files, query=query, frequency=frequency)
    base_dirname = os.path.basename(dir)
    cleanup_old_files(
        dir,
        filename_prefix=f"##{base_dirname}-",
        max_count_str=max_files,
        pattern="xlsx"
    )

#    log_message(progress_log_path, f"dir: {dir}, Documents exported: {len(document_data)}")
#    print(f"Exported Excel file: {excel_file}")

import time

def cache_is_fresh(all_dir: str, max_age_seconds: int = 3600) -> bool:
    """Prüft, ob der Cache im .all-Verzeichnis jünger als max_age_seconds ist."""
    ts_file = os.path.join(all_dir, "##cache.timestamp")
    try:
        if not os.path.exists(ts_file):
            return False
        mtime = os.path.getmtime(ts_file)
        age = time.time() - mtime
        if age < max_age_seconds:
            message(f"🕒 Cache im {all_dir} ist {int(age)}s alt – überspringe Neuaufbau.", "both")
            return True
        else:
            message(f"♻️ Cache im {all_dir} ist zu alt ({int(age)}s) – wird neu aufgebaut.", "both")
            return False
    except Exception as e:
        message(f"⚠️ Fehler bei Cacheprüfung: {e}", "both")
        return False


def update_cache_timestamp(all_dir: str):
    """Setzt oder aktualisiert den Timestamp des Cache."""
    ts_file = os.path.join(all_dir, "##cache.timestamp")
    try:
        with open(ts_file, "w") as f:
            f.write(str(int(time.time())))
        message(f"🗓️ Cache-Timestamp aktualisiert ({ts_file})", "log")
    except Exception as e:
        message(f"⚠️ Fehler beim Setzen des Cache-Timestamps: {e}", "both")


from pypaperless.models.generators.page import Page

async def safe_document_iterator(paperless):
    page_iter = aiter(paperless.documents.pages())

    while True:
        try:
            page: Page = await retry_async(lambda: anext(page_iter), desc="Lade Dokument-Seite")
        except StopAsyncIteration:
            break
        except Exception as e:
            message(f"Fehler beim Laden einer Seite: {e}")
            break

        for doc in page.items:
            yield doc

async def build_all_cache(paperless, export_dir, log_path=None):
    def log(msg):
        if log_path:
            message( msg)

    all_dir = os.path.join(export_dir, ".all")
    os.makedirs(all_dir, exist_ok=True)

    if cache_is_fresh(all_dir):
        #message("✅ Cache ist aktuell – überspringe Aufbau", "both")
        return

    # Schneller Count über .all()
    doc_ids = await retry_async(lambda: paperless.documents.all(), desc="Zähle Dokumente")
    total = len(doc_ids)

    done = 0
    cached = 0
    bar = tqdm(total=total, desc="Dokumente cachen: 0✓ / 0↓", unit="doc")

    async for doc in safe_document_iterator(paperless):
        try:
            sanitized_title = sanitize_filename(doc.title)
            pdf_filename = f"{doc.id}--{sanitized_title}.pdf"
            json_filename = f"{doc.id}--{sanitized_title}.json"
            pdf_path = os.path.join(all_dir, pdf_filename)
            json_path = os.path.join(all_dir, json_filename)

            updated = False

            if not os.path.exists(pdf_path):
                ok = await export_pdf(doc, working_dir=all_dir)
                updated = updated or ok

            if not os.path.exists(json_path):
                try:
                    export_json(paperless=paperless, doc=doc, working_dir=all_dir)
                    updated = True
                except Exception as e:
                    message(f"JSON-Export fehlgeschlagen für Doc {doc.id}: {e}", target="log", level="warn")

            if updated:
                done += 1
            else:
                cached += 1

            bar.update(1)
            bar.set_description(f"Dokumente cachen: {cached}✓ / {done}↓")

            # minimaler Cooldown, um Verbindungsabbrüche zu reduzieren
            await asyncio.sleep(0.05)

        except Exception as e:
            # Fängt alles, damit ein einzelnes Dokument nicht den gesamten Lauf killt
            message(f"Fehler bei Doc {getattr(doc,'id','?')}: {e}", target="log", level="warn")
            bar.update(1)

    update_cache_timestamp(all_dir)


def extract_doc_id(filename):
    """Extrahiere die Dokument-ID aus einem Dateinamen wie '874--irgendwas.pdf'."""
    try:
        return int(filename.split('--', 1)[0])
    except (ValueError, IndexError):
        return None

async def cleanup_all_dir(paperless, all_dir, log_path=None):
    message( "Bereinige .all-Verzeichnis...")

    # Aktuelle Dokument-IDs abrufen
    valid_doc_ids = set(await retry_async(lambda: paperless.documents.all(), desc="Lade gültige Dokument-IDs"))

    removed_files = 0
    for filename in os.listdir(all_dir):
        if not (filename.endswith(".pdf") or filename.endswith(".json")):
            continue

        doc_id = extract_doc_id(filename)
        if doc_id is None or doc_id not in valid_doc_ids:
            file_path = os.path.join(all_dir, filename)
            try:
                os.remove(file_path)
                removed_files += 1
                message(f"Entfernt: {filename}")
            except Exception as e:
                message(message=f"Fehler beim Löschen von {filename}: {e}")

    message(text=f"Bereinigung abgeschlossen: {removed_files} Datei(en) entfernt.")

def is_remote_newer(remote_modified_str, local_path):
    # Vergleiche Timestamps
    remote_time = datetime.datetime.fromisoformat(remote_modified_str)
    local_time = datetime.datetime.fromtimestamp(os.path.getmtime(local_path))
    return remote_time > local_time

async def main():
    print_program_header()
    print_separator('=', 0.75)

    script_name = get_script_name()
    config = load_config_from_script()

    # 🔍 Pflichtfelder prüfen
    required = {
        "API": ["url", "token"],
        "Export": ["directory"],
        "Log": ["log_file", "max_files"]
    }

    missing = []
    for section, keys in required.items():
        if not config.has_section(section):
            missing.append(f"[{section}] fehlt")
            continue
        for key in keys:
            if not config.has_option(section, key) or not config.get(section, key).strip():
                missing.append(f"{section}.{key} fehlt oder ist leer")

    if missing:
        print("\n❌ Fehler in der Konfigurationsdatei:")
        for m in missing:
            print(f"   - {m}")
        print("\n💡 Bitte prüfe deine .ini-Datei und ergänze die fehlenden Angaben.")
        sys.exit(1)

    # ✅ Nur wenn alles ok ist: Konfigurationswerte lesen
    export_dir = config.get("Export", "directory")
    api_url = config.get("API", "url")
    api_token = config.get("API", "token")
    log_dir = config.get("Log", "log_file")
    max_files = config.get("Log", "max_files")

    # Log-Dateien initialisieren
    prepare_logging(log_dir, script_name, max_files)

    # Beispielverwendung nach Initialisierung:
    # set_log_path("pfad/zur/logdatei.log")
    # message("Export gestartet", target="both")
    # message("Dokument konnte nicht geladen werden", level="warn")
    # message("Verarbeite Dokument 17", inline=True)

# Setze das Arbeitsverzeichnis auf das Verzeichnis, in dem das Skript gespeichert ist
    script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    os.chdir(script_dir)

# Setze Locale früh im Programm (fallback auf Standard-Locale)
    try:
        locale.setlocale(locale.LC_ALL, 'de_DE.UTF-8')
    except locale.Error:
        try:
            locale.setlocale(locale.LC_ALL, '')
        except:
            print("⚠️ Locale konnte nicht gesetzt werden – Formatierung ggf. fehlerhaft.")

    message(f"Log in to {api_url} with {mask_secret(api_token)} ..", target="both")
    paperless = Paperless(api_url, api_token)
    await retry_async(lambda: paperless.initialize(), desc="Paperless-Login")
    
    message("Logged in", target="both")
    message(f"Export to {export_dir}", target="both")

    # do something
    meta = await fetch_paperless_meta(paperless)
    # Zugriff auf ein Element
    #print(meta["correspondents"][3].name)
    #print(meta["tags"][1].name)
    #print(meta["storage_paths"][2].name)

    await build_all_cache(paperless, export_dir)
    await cleanup_all_dir(paperless, all_dir=os.path.join(export_dir, ".all"))

    excluded_dirs = {"@eaDir", ".all"}
    try:
        for root, dirs, files in os.walk(export_dir):
          query_value = os.path.basename(root)
          dirs[:] = [d for d in dirs if d not in excluded_dirs]
          if root == export_dir:
            continue

          config_mtime = 0  # oder datetime.min.timestamp()
          if '##config.ini' in files:
              config_path = os.path.join(root, '##config.ini')
              config = configparser.ConfigParser()
           #   config.read(config_path)
              config=load_config(config_path=config_path)
              config_mtime = os.path.getmtime(config_path)

          if 'DATA' in config and 'query' in config['DATA']:
              query_value = config['DATA']['query']

          if 'EXPORT' in config and 'frequency' in config['EXPORT']:
              frequency = config['EXPORT']['frequency']
          else:
              frequency = 'daily'

          should_run, reason = should_export(root, frequency, config_mtime)
          message(f"{root} : {query_value} -> ({reason})", target="both")
          if should_run:
              #print_separator('#')           # #######...
              #print_separator('##')          # ## ## ## ...
              #print_separator('=')           # ==========...
              #print_separator('·', 0.5)      # 50% der Breite
              print_separator('=', 0.75)      # 50% der Breite
              #print(f"\n{root} {query_value} -> Export ({reason})")
              await exportThem(paperless=paperless, dir=root, query=query_value, max_files=max_files,frequency=frequency, api_url=api_url)
          else:
              #print(f"\n{root} {query_value} -> NOexport ({reason})")
              print_separator('-', 0.75)      # 50% der Breite

    except Exception as e:
        message(f"Error: {str(e)}", target="both")
        raise
    finally:
        if paperless:
            await paperless.close()
        #finalize_log(progress_log_path, final_log_path)

asyncio.run(main())
